from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver import Remote
from bs4 import BeautifulSoup
from bs4 import element
from openpyxl.worksheet import worksheet
from openpyxl import load_workbook
from openpyxl import Workbook

from typing import NamedTuple
from os import environ


class UserConf_NamedTuple(NamedTuple):
    id: str
    password: str


def load_conf() -> UserConf_NamedTuple:
    return UserConf_NamedTuple(
        id=environ["LOGIN_ID"], password=environ["LOGIN_PASSWORD"]
    )


def set_up_chrome_driver() -> Remote:
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option("useAutomationExtension", False)
    chrome_driver = Remote(
        command_executor=environ["SELENIUM_URL"], options=chrome_options
    )
    # set timeout for find elements or for complete commands
    chrome_driver.implicitly_wait(10)
    return chrome_driver


def login_dream_campas(chrome_driver: Remote, login_info: UserConf_NamedTuple):
    chrome_driver.get("https://www2.st.kagawa-u.ac.jp/Portal/")
    chrome_driver.find_element(By.ID, "txtID").send_keys(login_info.id)
    chrome_driver.find_element(By.ID, "txtPassWord").send_keys(login_info.password)
    chrome_driver.find_element(By.ID, "btnLogIn").click()


def get_registlist_atags(chrome_driver: Remote) -> list[element.Tag]:
    chrome_driver.get(
        "https://www2.st.kagawa-u.ac.jp/Portal/StudentApp/Regist/RegistList.aspx"
    )
    html = chrome_driver.page_source.encode("utf-8")
    soup = BeautifulSoup(html, "html.parser")
    a_tags: element.ResultSet[element.Tag] = soup.find_all("a", href=True)
    # filter a_tag that has href linking toward the syllabus page
    return [a_tag for a_tag in a_tags if "Portal/Public/Syllabus" in str(a_tag["href"])]


def get_syllabus_url(a_tag: element.Tag) -> str:
    href = str(a_tag["href"])
    return href


def get_lesson_name(a_tag: element.Tag) -> str:
    lesson_name = str(a_tag.string)
    return lesson_name


def get_textbook(chrome_driver: Remote, a_tag: element.Tag) -> str:
    syllabus_url = get_syllabus_url(a_tag)
    chrome_driver.get(syllabus_url)
    # 「教科書・参考書等」の内容が入っている span tag の中身
    text_description = chrome_driver.find_element(
        By.ID, "ctl00_phContents_ucSylContents_cateRequiredTexts_lblNormal"
    ).text
    return text_description


def create_lesson_name_and_textbook_dict(
    chrome_driver: Remote, registlist_a_tag_list: list[element.Tag]
) -> dict[str, str]:
    lesson_and_textbook: dict[str, str] = {}
    for registlist_a_tag in registlist_a_tag_list:
        lesson_name = get_lesson_name(registlist_a_tag)
        textbook = get_textbook(chrome_driver, registlist_a_tag)
        lesson_and_textbook[lesson_name] = textbook
    return lesson_and_textbook


def excel_insert_line(
    sheet: worksheet.Worksheet, line: list, start_row: int, start_column: int
) -> None:
    for elem_idx, line_elem in enumerate(line):
        sheet.cell(row=start_row, column=start_column + elem_idx, value=line_elem)


def save_to_excel(lesson_and_textbooks: dict[str, str]) -> None:
    EXCEL_PATH = environ["EXCEL_PATH"]
    SHEET_NAME = "Sheet"
    try:
        book = load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        book = Workbook()
    try:
        sheet = book[SHEET_NAME]
    except KeyError:
        sheet = book.create_sheet(SHEET_NAME)
    legends = ["講義名", "教科書・参考書等"]
    excel_insert_line(sheet=sheet, line=legends, start_row=1, start_column=1)
    for idx, (lesson_name, textbook_descriptions) in enumerate(
        lesson_and_textbooks.items()
    ):
        excel_insert_line(
            sheet=sheet,
            line=[lesson_name, textbook_descriptions],
            start_row=2 + idx,
            start_column=1,
        )
    book.save(EXCEL_PATH)


if __name__ == "__main__":
    chrome_driver = set_up_chrome_driver()
    login_dream_campas(chrome_driver, load_conf())
    a_tags = get_registlist_atags(chrome_driver)
    lesson_and_textbook = create_lesson_name_and_textbook_dict(chrome_driver, a_tags)
    chrome_driver.quit()
    save_to_excel(lesson_and_textbook)
